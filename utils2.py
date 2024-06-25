import openpyxl
import openpyxl.workbook
from openpyxl.styles import NamedStyle
import glob

date_style = NamedStyle(name='custom_datetime', number_format='dd-mmm-yyyy')

def get_sorted_merged_cells(wb, sheet):
    merged_cells_list = []
    for merged_cell in wb[sheet].merged_cells.ranges:
        merged_cells_list.append(merged_cell)
    return sorted(merged_cells_list, key=lambda x: (x.min_row, x.min_col))

def get_only_non_header_merged_cells(wb, sheet, header_num_rows):
    merged_cells = get_sorted_merged_cells(wb, sheet)
    return [merged_cell for merged_cell in merged_cells if merged_cell.min_row > header_num_rows]

# function to check current cell is part of merged cell or not
def is_merged_cell(merged_cells, cell):
    for index in range(len(merged_cells)):
        if cell.row >= merged_cells[index].min_row and cell.row <= merged_cells[index].max_row and cell.column >= merged_cells[index].min_col and cell.column <= merged_cells[index].max_col:
            return True, index
    return False, -1

def find_headers(xlsx_files):
    if len(xlsx_files)>0:
        xlsx_file = xlsx_files[0]
        wb = openpyxl.load_workbook(xlsx_file)
        other_wbs = [ openpyxl.load_workbook(xlsx_file) for xlsx_file in xlsx_files[1:] ]
        sheets = wb.sheetnames
        print(sheets)

        sheets_headers_num_rows = {sheet:-1 for sheet in sheets}
        for sheet in sheets:
            merged_cells = get_sorted_merged_cells(wb, sheet)
            print(merged_cells)
            max_row, max_col = wb[sheet].max_row, wb[sheet].max_column
            header_num_rows = -1
            flag = False
            for i in range(1, max_row+1):
                for j in range(1, max_col+1):
                    cell = wb[sheet].cell(row=i, column=j)
                    for other_wb in other_wbs:
                        other_cell = other_wb[sheet].cell(row=i, column=j)
                        if cell.value != other_cell.value:
                            flag = True
                            break

                    is_merged, index = is_merged_cell(merged_cells, cell)
                    if is_merged and merged_cells[index].min_row == cell.row and merged_cells[index].min_col == cell.column:
                        print(f'xxxxxxxxxxxxxxxxxxxxxxxxx Content of merged cell {merged_cells[index]} is {cell.value}')
                    elif is_merged:
                        continue
                    else:
                        print(f'------------------------- Content of cell {cell.coordinate} is {cell.value}')
                    if flag:
                        break
                if flag:
                    break
                header_num_rows = i
            sheets_headers_num_rows[sheet] = header_num_rows
        print(sheets_headers_num_rows)
        return sheets_headers_num_rows
    
def convert_to_coord(row,col):
    return f'{openpyxl.utils.get_column_letter(col)}{row}'

def get_merged_cell_str(cell1, cell2):
    return f'{cell1}:{cell2}'

def get_merged_cell_from_coord(row1, col1, row2, col2):
    return get_merged_cell_str(convert_to_coord(row1, col1), convert_to_coord(row2, col2))

def get_row_col_from_coord(coord):
    return openpyxl.utils.cell.coordinate_to_tuple(coord)
    
def merge_excels(excel_files, output_file="Merged_Report.xlsx"):
    sheets_headers_num_rows = find_headers(excel_files)
    wb = openpyxl.Workbook()
    # remove default sheet
    wb.remove(wb.active)
    wb_excel = openpyxl.load_workbook(excel_files[0])
    for sheet in wb_excel.sheetnames:
        ws = wb.create_sheet(sheet)
        ws_excel = wb_excel[sheet]
        for i in range(1, sheets_headers_num_rows[sheet]+1):
            for j in range(1, ws_excel.max_column+1):
                ws.cell(row=i, column=j, value=ws_excel.cell(row=i, column=j).value)
        merged_cells = get_sorted_merged_cells(wb_excel, sheet)
        # filter merged cells which are part of header
        merged_cells = [merged_cell for merged_cell in merged_cells if merged_cell.min_row <= sheets_headers_num_rows[sheet]]
        for merged_cell in merged_cells:
            ws.merge_cells(merged_cell.coord)
        ws.freeze_panes = ws[convert_to_coord(sheets_headers_num_rows[sheet]+1, 1)]

    # # copy all cells from first sheet of first excel file to new workbook
    ws = wb[wb_excel.sheetnames[0]]
    for i in range(sheets_headers_num_rows[wb_excel.sheetnames[0]]+1, wb_excel[wb_excel.sheetnames[0]].max_row+1):
        for j in range(1, wb_excel[wb_excel.sheetnames[0]].max_column+1):
            ws.cell(row=i, column=j, value=wb_excel[wb_excel.sheetnames[0]].cell(row=i, column=j).value)

    merged_cells = get_sorted_merged_cells(wb_excel, wb_excel.sheetnames[0])
    for merged_cell in merged_cells:
        ws.merge_cells(merged_cell.coord)

    for sheet in wb_excel.sheetnames[1:]:
        ws = wb[sheet]
        starting_row = sheets_headers_num_rows[sheet]+1
        for wb_file in excel_files:
            print("----------------------------------------------"+wb_file+"----"+sheet+"----------------------------------------------")
            wb_excel = openpyxl.load_workbook(wb_file)
            ws_excel = wb_excel[sheet]
            num_rows = ws_excel.max_row - sheets_headers_num_rows[sheet]
            for i in range(starting_row, starting_row + num_rows):
                for j in range(1, ws_excel.max_column+1):
                    ws_excel_cell_value = ws_excel.cell(row=i - starting_row + 1 + sheets_headers_num_rows[sheet], column=j).value
                    # print(f'Content of cell {i} and {j} is {ws_excel_cell_value}')
                    ws.cell(row=i, column=j, value=ws_excel_cell_value)
            merged_cells = get_only_non_header_merged_cells(wb_excel, sheet, sheets_headers_num_rows[sheet])
            for merged_cell in merged_cells:
                start_coords, end_coords = merged_cell.coord.split(':')
                start_coords = get_row_col_from_coord(start_coords)
                end_coords = get_row_col_from_coord(end_coords)
                # print(f'Current coords are {start_coords} and {end_coords}')
                new_start_coords = (starting_row + start_coords[0] - 1 - sheets_headers_num_rows[sheet], start_coords[1])
                new_end_coords = (starting_row + end_coords[0] - 1 - sheets_headers_num_rows[sheet], end_coords[1])
                # print(f'New coords are {new_start_coords} and {new_end_coords}')
                ws.merge_cells(get_merged_cell_from_coord(new_start_coords[0], new_start_coords[1], new_end_coords[0], new_end_coords[1]))
            starting_row = ws.max_row+1

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for i in range(1, ws.max_column+1):
            max_length = 0
            cell_value = ""
            merged_cells = get_sorted_merged_cells(wb, sheet)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    # try:
                        # if cell is merged cell then get value of first cell
                    # print(f'aaaa Content of cell {cell.coordinate} is {cell.value} {is_merged_cell(get_sorted_merged_cells(wb, sheet), cell)}')

                    if is_merged_cell(merged_cells, cell)[0]:
                        # print(f'Content of merged cell {cell.coordinate} is {cell.value} {is_merged_cell(merged_cells, cell)[1]}')
                        merged_cell_start, merged_cell_end = merged_cells[is_merged_cell(merged_cells, cell)[1]].coord.split(':')
                        merged_cell_start = get_row_col_from_coord(merged_cell_start)
                        merged_cell_end = get_row_col_from_coord(merged_cell_end)
                        # print(merged_cell_start, merged_cell_end)
                        
                        if merged_cell_start[1] == merged_cell_end[1]:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                                cell_value = cell.value
                                print(("mmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmm" + str(max_length) + " " + str(cell.value)))
                                if len(str(cell.value)) > 100:
                                    print(f'bbbbbb    Content of merged cell {cell.coordinate} is {cell.value}')  
                                    print(f"{merged_cell_start[0]} {merged_cell_end[0]} {merged_cell_start[1]} {merged_cell_end[1]}")
                                    raise Exception("s,fjvb")
                            
                    elif len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                        cell_value = cell.value
                        print(("nnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnn" + str(max_length) + " " + str(cell.value)))
                    # except Exception as e:
                    #     print("eeeee "+str(e))
            print(f'Column {i} has max length {max_length} {cell_value}')
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_length + 1
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].number_format = '0.00'
    wb.save(output_file)

    
    
# find_headers(['1.xlsx', '2.xlsx', '3.xlsx', '4.xlsx'])
if __name__ == "__main__":
# all xlsx files in data/type-1 folder
    xlsx_files = glob.glob("data/type-2/*.xlsx")

    merge_excels(['1.xlsx', '2.xlsx', '3.xlsx', '4.xlsx'])
# merge_excels(xlsx_files)