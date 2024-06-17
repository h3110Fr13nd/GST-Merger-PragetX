from openpyxl import load_workbook
import pandas as pd
import datetime


month_map = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def merge_excels_by_sheet_name(excel_files=None, sheet_name=None):
    df = None
    for file in excel_files:
        excel_file = pd.ExcelFile(file)
        if df is None:
            df = excel_file.parse(sheet_name, header=[0, 1, 2])
        else:
            df = pd.concat([df, excel_file.parse(sheet_name, header=[0, 1, 2])])
    return df


def get_min_max_daterange(files=None):
    min_date = (9999, 13)
    max_date = (0, 0)
    for file in files:
        wb = load_workbook(file)
        sheet = wb.active
        cell_value = sheet["B4"].value
        cell_value = cell_value.split("-")
        month = cell_value[0].strip().lower()
        year = int(cell_value[1].strip())
        month_number = month_map[month]
        if (year, month_number) < min_date:
            min_date = (year, month_number)
        if (year, month_number) > max_date:
            max_date = (year, month_number)
    print("min_date", min_date)
    return min_date, max_date


def get_min_max_date_string(min_date=None, max_date=None):
    min_month = list(month_map.keys())[
        list(month_map.values()).index(min_date[1])
    ].capitalize()
    max_month = list(month_map.keys())[
        list(month_map.values()).index(max_date[1])
    ].capitalize()
    return f"{min_month} {min_date[0]} - {max_month} {max_date[0]}"


def merge_and_save(files=None, output_file="GST-Report(Yearly).xlsx"):
    writer = pd.ExcelWriter(output_file, datetime_format='dd-mmm-yyyy')
    first_file = files[0]
    excel_file = pd.ExcelFile(first_file)
    workbook = writer.book
    cell_format = workbook.add_format()
    cell_format.set_align("right")
    sheets = excel_file.sheet_names
    sheet1 = excel_file.parse(sheets[0])
    sheet1.to_excel(writer, sheet_name=sheets[0], index=False, merge_cells=True)
    writer.sheets[sheets[0]].write(
        "B4", get_min_max_date_string(*get_min_max_daterange(files))
    )
    # today's date in dd-mm-yyyy format
    writer.sheets[sheets[0]].write("B5", datetime.datetime.now().strftime("%d-%m-%Y"))
    for idx, col in enumerate(sheet1):  # loop through all columns
        series = sheet1[col]
        max_in_series = series.astype(str).map(len).max()
        max_in_series = 0 if pd.isnull(max_in_series) else max_in_series
        max_len = min(max(max_in_series, len(str(series.name))) + 3, 50)
        writer.sheets[sheets[0]].set_column(idx, idx, max_len)
    # merge A1 and B1 with value of A1
    writer.sheets[sheets[0]].merge_range(
        "A1:B1",
        sheet1.columns[0],
        writer.book.add_format({"align": "center", "valign": "vcenter"}),
    )
    for sheet in sheets[1:]:
        merged_data = merge_excels_by_sheet_name(
            excel_files=files, sheet_name=sheet
        ).reset_index(drop=True)
        merged_data.columns = merged_data.columns.map(
            lambda x: (x[0], "", x[2]) if isinstance(x, tuple) else x
        )
        merged_data.to_excel(
            writer, header=[0, 1, 2], sheet_name=sheet, freeze_panes=(3, 0)
        )
        if merged_data.shape[0] == 0:
            print("sheet", sheet, "has no rows")
        for idx, col in enumerate(merged_data):  # loop through all columns
            series = merged_data[col]
            max_in_series = series.astype(str).map(len).max()
            max_in_series = 0 if pd.isnull(max_in_series) else max_in_series
            max_len = min(max((max_in_series, len(str(series.name[2])))) + 3, 50)
            writer.sheets[sheet].set_column(idx + 1, idx + 1, max_len, cell_format)
        writer.sheets[sheet].set_row(3, None, None, {"hidden": True})
    writer.close()
